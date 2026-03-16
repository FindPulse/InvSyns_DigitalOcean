#!/usr/bin/env python3
"""
Ingest a combined WTD wheel offer Excel that contains multiple brands in ONE sheet.

Your requirement (Rohana-style):
- The script MUST pick the file from: D:\app\InvSync\Vendor_Files
- You pass ONLY the filename via --filename (no full path)
- as_of_date MUST be parsed ONLY from filename (if not found -> error)

This file pattern (confirmed from your upload):
- Sheet: "Brands"
- Columns:
    SKU  -> "Product Code"
    QTY  -> "Avail Qty"
- There are MANY product images in column A (ignore these)
- The brand separator/logo images are in the middle columns (not column A)
- Dolce appears in 2 sections (Dolce + Dolce Performance) -> both map to Dolce vendor

Loads into:
  invsync.vendors
  invsync.vendor_files (dedupe by sha256 per vendor)
  invsync.vendor_stock_raw
  invsync.inventory_normalized
"""

import argparse
import hashlib
import json
import os
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
load_dotenv()

import pandas as pd
import psycopg2
import psycopg2.extras

from openpyxl import load_workbook

# ----------------------------
# Config (Rohana-style)
# ----------------------------
SCHEMA = "invsync"
INVROOT= os.path.dirname(os.path.abspath(__file__))
VENDOR_FILES_DIR = os.path.join(INVROOT, "Vendor_Files")

DEFAULT_SHEET_NAME = "Brands"

# Confirmed column names in your file
SKU_COL = "Product Code"
QTY_COL = "Avail Qty"

# We will IGNORE images anchored to Column A (product images).
# Brand logos appear in other columns (middle), and those rows define sections.


# ----------------------------
# Helpers
# ----------------------------
def sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_date_from_filename_only(filename: str) -> date:
    """
    Must parse ONLY from filename. If not found -> raise.

    Supported examples:
      WTDwheelofferFebruary_10_2026.xlsx  (MonthName_DD_YYYY)
      Rohana Inventory 01-16-2026.xlsx    (MM-DD-YYYY)
      vendor_2026-02-10.xlsx              (YYYY-MM-DD)
    """
    fn = filename.lower()

    # MonthName_DD_YYYY (your file)
    m = re.search(
        r"(january|february|march|april|may|june|july|august|september|october|november|december)"
        r"[_\-\s]+(\d{1,2})[_\-\s]+(\d{4})",
        fn,
    )
    if m:
        month_name, dd, yyyy = m.groups()
        month_map = {
            "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
            "july": 7, "august": 8, "september": 9, "october": 10, "november": 11, "december": 12
        }
        return date(int(yyyy), month_map[month_name], int(dd))

    # MM-DD-YYYY
    m = re.search(r"(\d{1,2})-(\d{1,2})-(\d{4})", filename)
    if m:
        mm, dd, yyyy = map(int, m.groups())
        return date(yyyy, mm, dd)

    # YYYY-MM-DD
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", filename)
    if m:
        yyyy, mm, dd = map(int, m.groups())
        return date(yyyy, mm, dd)

    raise RuntimeError(
        f"Could not parse as_of_date from filename '{filename}'. "
        f"Rename file to include a date like 'February_10_2026' or '02-10-2026' or '2026-02-10'."
    )


def get_conn():
    host = os.environ.get("PGHOST")
    db = os.environ.get("PGDATABASE")
    user = os.environ.get("PGUSER")
    pw = os.environ.get("PGPASSWORD")
    port = int(os.environ.get("PGPORT", "5432"))
    sslmode = os.environ.get("PGSSLMODE", "require")

    if not all([host, db, user, pw]):
        raise RuntimeError(
            "Missing PG env vars. Need PGHOST, PGDATABASE, PGUSER, PGPASSWORD "
            "(and optionally PGPORT, PGSSLMODE)."
        )

    return psycopg2.connect(
        host=host,
        port=port,
        dbname=db,
        user=user,
        password=pw,
        sslmode=sslmode,
    )


def ensure_vendor(cur, vendor_code: str, vendor_name: str) -> int:
    cur.execute(
        f"""
        INSERT INTO {SCHEMA}.vendors (vendor_code, vendor_name)
        VALUES (%s, %s)
        ON CONFLICT (vendor_code) DO UPDATE
          SET vendor_name = EXCLUDED.vendor_name,
              updated_at = NOW()
        RETURNING vendor_id;
        """,
        (vendor_code, vendor_name),
    )
    return int(cur.fetchone()[0])


def insert_vendor_file(
    cur,
    vendor_id: int,
    file_path: str,
    file_hash: str,
    source_email: Optional[str],
    subject: Optional[str],
) -> Tuple[int, str]:
    filename = os.path.basename(file_path)
    size_bytes = os.path.getsize(file_path)

    cur.execute(
        f"""
        INSERT INTO {SCHEMA}.vendor_files
          (vendor_id, source_email, subject, filename, file_hash_sha256, file_size_bytes, received_at, status)
        VALUES
          (%s, %s, %s, %s, %s, %s, NOW(), 'received')
        ON CONFLICT (vendor_id, file_hash_sha256) DO UPDATE
          SET processed_at = COALESCE({SCHEMA}.vendor_files.processed_at, NOW()),
              status = {SCHEMA}.vendor_files.status
        RETURNING file_id, status::text;
        """,
        (vendor_id, source_email, subject, filename, file_hash, size_bytes),
    )
    file_id, status = cur.fetchone()
    return int(file_id), str(status)


def to_int_qty(x: Any) -> int:
    if x is None or pd.isna(x):
        return 0
    if isinstance(x, (int, float)):
        q = int(x)
        return q if q >= 0 else 0
    s = str(x).strip()
    if s == "":
        return 0
    s = re.sub(r"[^\d\-]", "", s)
    try:
        q = int(s)
        return q if q >= 0 else 0
    except Exception:
        return 0


# ----------------------------
# Brand marker detection
# ----------------------------
def detect_non_product_image_rows(ws) -> List[int]:
    """
    Returns Excel row numbers (1-based) that contain an embedded image whose anchor column != 1.
    Column 1 (A) images are product photos -> ignored.
    """
    rows = set()
    images = getattr(ws, "_images", []) or []
    for img in images:
        try:
            a = img.anchor
            r0 = getattr(getattr(a, "_from", None), "row", None)  # 0-based
            c0 = getattr(getattr(a, "_from", None), "col", None)  # 0-based
            if r0 is None or c0 is None:
                continue
            excel_row = int(r0) + 1
            excel_col = int(c0) + 1
            if excel_col == 1:
                continue  # product photo
            rows.add(excel_row)
        except Exception:
            continue
    return sorted(rows)


def try_ocr_brand_markers(ws, marker_rows: List[int]) -> Dict[int, str]:
    """
    Optional: OCR brand logos to map marker row -> brand code.
    If OCR dependencies are missing, returns {} and we will fallback to order-based mapping.
    """
    try:
        from PIL import Image, ImageEnhance
        import pytesseract
        import io
    except Exception:
        return {}

    row_to_brand: Dict[int, str] = {}
    images = getattr(ws, "_images", []) or []

    # Build quick index of images by row, but only for non-column-A markers
    for img in images:
        try:
            a = img.anchor
            r0 = getattr(getattr(a, "_from", None), "row", None)
            c0 = getattr(getattr(a, "_from", None), "col", None)
            if r0 is None or c0 is None:
                continue
            excel_row = int(r0) + 1
            excel_col = int(c0) + 1
            if excel_col == 1:
                continue
            if excel_row not in marker_rows:
                continue

            b = img._data()
            im = Image.open(io.BytesIO(b)).convert("L")
            im = ImageEnhance.Contrast(im).enhance(3)
            im = im.resize((im.size[0] * 3, im.size[1] * 3))
            txt = pytesseract.image_to_string(im).lower()

            # Map logo text -> vendor code
            if "dolce" in txt:
                row_to_brand[excel_row] = "DOL"
            elif "katana" in txt:
                row_to_brand[excel_row] = "KAT"
            elif "lock" in txt and "off" in txt:
                row_to_brand[excel_row] = "LOR"
            elif "wtd" in txt or "wholesale tire" in txt:
                row_to_brand[excel_row] = "WTD"
        except Exception:
            continue

    return row_to_brand


def assign_brand_by_marker_order(marker_rows: List[int]) -> Dict[int, str]:
    """
    Fallback mapping if OCR is not available:
    - Ignore the first markers that are likely header/other brand (WTD + Lock),
      and map the remaining markers by order: Dolce, Dolce, Katana.

    This matches the structure you uploaded:
      Row 3  : WTD (header)
      Row 4  : Lock
      Row 786: Dolce
      Row 1271: Dolce Performance
      Row 1512: Katana

    If file structure changes, OCR is safer.
    """
    mapping: Dict[int, str] = {}
    if len(marker_rows) < 3:
        return mapping

    # Heuristic: last 3 markers are Dolce / Dolce / Katana (as in your file)
    tail = marker_rows[-3:]
    if len(tail) == 3:
        mapping[tail[0]] = "DOL"
        mapping[tail[1]] = "DOL"
        mapping[tail[2]] = "KAT"
    return mapping


# ----------------------------
# Main
# ----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--filename",
        required=True,
        help="Excel filename located in D:\\app\\InvSync\\Vendor_Files (e.g. WTDwheelofferFebruary_10_2026.xlsx)",
    )
    ap.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="Sheet name (default: Brands)")

    # Vendor overrides (defaults DOL/KAT as requested)
    ap.add_argument("--dolce-code", default="DOL", help="Vendor code for Dolce (default: DOL)")
    ap.add_argument("--dolce-name", default="Dolce", help="Vendor name for Dolce (default: Dolce)")
    ap.add_argument("--katana-code", default="KAT", help="Vendor code for Katana (default: KAT)")
    ap.add_argument("--katana-name", default="Katana", help="Vendor name for Katana (default: Katana)")

    # If you ever want to load Lock too, set this false and add lock vendor below
    ap.add_argument("--only-dolce-katana", default="true", help="true/false (default: true)")

    ap.add_argument("--source-email", default=None)
    ap.add_argument("--subject", default=None)
    args = ap.parse_args()

    file_path = os.path.join(VENDOR_FILES_DIR, args.filename)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found in Vendor_Files: {file_path}")

    # as_of_date MUST come from filename only
    as_of = parse_date_from_filename_only(args.filename)

    only_dk = str(args.only_dolce_katana).strip().lower() in ("1", "true", "yes", "y")

    file_hash = sha256_file(file_path)

    # Load workbook for image markers
    wb = load_workbook(file_path, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{args.sheet}' not found. Found: {wb.sheetnames}")
    ws = wb[args.sheet]

    marker_rows = detect_non_product_image_rows(ws)
    if not marker_rows:
        raise RuntimeError(
            "No non-column-A images found to use as brand markers. "
            "If the brand logos are not stored as embedded images, we need another rule."
        )

    # Try OCR mapping first; fallback to order mapping
    row_to_brand = try_ocr_brand_markers(ws, marker_rows)
    if not row_to_brand:
        row_to_brand = assign_brand_by_marker_order(marker_rows)

    if not row_to_brand:
        raise RuntimeError(
            f"Could not map brand marker rows to brands. Marker rows found: {marker_rows}. "
            f"Install OCR deps (Pillow + pytesseract + Tesseract) OR we can hard-map rows."
        )

    # Read tabular data (pandas ignores images automatically)
    df = pd.read_excel(file_path, sheet_name=args.sheet, header=0)
    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in [SKU_COL, QTY_COL] if c not in df.columns]
    if missing:
        raise RuntimeError(f"Missing required columns {missing}. Found: {list(df.columns)}")

    # pandas row index 0 corresponds to Excel row 2 (header is row 1)
    def df_index_to_excel_row(i: int) -> int:
        return int(i) + 2

    # Walk rows, track current brand based on latest marker row seen
    current_brand: Optional[str] = None

    # Sort markers so we can advance efficiently
    sorted_marker_rows = sorted(row_to_brand.keys())
    marker_ptr = 0

    raw_by_brand: Dict[str, List[Tuple[int, int, Optional[str], Optional[str], Dict[str, Any], bool, Optional[str]]]] = {}
    norm_by_brand: Dict[str, List[Tuple[int, int, date, str, int, Optional[int]]]] = {}

    for i, row in df.iterrows():
        excel_row = df_index_to_excel_row(i)

        # Advance marker pointer and set current brand when we pass marker rows
        while marker_ptr < len(sorted_marker_rows) and excel_row >= sorted_marker_rows[marker_ptr]:
            b = row_to_brand[sorted_marker_rows[marker_ptr]]
            # Ignore non-product header markers if OCR detected them
            if b in ("DOL", "KAT", "LOR"):
                current_brand = b
            marker_ptr += 1

        if current_brand is None:
            continue
        if only_dk and current_brand not in ("DOL", "KAT"):
            continue

        raw_sku = None if pd.isna(row.get(SKU_COL)) else str(row.get(SKU_COL)).strip()
        if not raw_sku:
            continue  # skip blank rows

        qty_val = row.get(QTY_COL)
        qty = to_int_qty(qty_val)
        raw_qty_str = None if pd.isna(qty_val) else str(qty_val).strip()

        payload: Dict[str, Any] = {}
        for c in df.columns:
            v = row.get(c)
            payload[c] = None if pd.isna(v) else (v.item() if hasattr(v, "item") else v)

        raw_by_brand.setdefault(current_brand, [])
        norm_by_brand.setdefault(current_brand, [])

        parsed_ok = True
        parse_error = None

        raw_by_brand[current_brand].append((0, excel_row, raw_sku, raw_qty_str, payload, parsed_ok, parse_error))
        norm_by_brand[current_brand].append((0, 0, as_of, raw_sku, qty, excel_row))

    # Vendor config with overrides
    vendor_cfg = {
        "DOL": (args.dolce_code, args.dolce_name),
        "KAT": (args.katana_code, args.katana_name),
        "LOR": ("LOR", "Lock Off-Road"),  # only used if only_dolce_katana=false
    }

    with get_conn() as conn:
        conn.autocommit = False
        with conn.cursor() as cur:
            vendor_ids: Dict[str, int] = {}
            file_ids: Dict[str, int] = {}

            # Ensure vendors + vendor_files per brand
            for brand in raw_by_brand.keys():
                vcode, vname = vendor_cfg[brand]
                vid = ensure_vendor(cur, vcode, vname)
                vendor_ids[brand] = vid

                fid, _ = insert_vendor_file(cur, vid, file_path, file_hash, args.source_email, args.subject)
                file_ids[brand] = fid

            # Insert per brand
            for brand in raw_by_brand.keys():
                vid = vendor_ids[brand]
                fid = file_ids[brand]

                raw_rows2 = [
                    (fid, row_num, sku, raw_qty, json.dumps(payload), ok, err)
                    for (_, row_num, sku, raw_qty, payload, ok, err) in raw_by_brand[brand]
                ]
                norm_rows2 = [
                    (fid, vid, as_of, sku, qty, src_row)
                    for (_, _, as_of_dt, sku, qty, src_row) in norm_by_brand[brand]
                    for as_of in [as_of_dt]
                ]

                if raw_rows2:
                    psycopg2.extras.execute_values(
                        cur,
                        f"""
                        INSERT INTO {SCHEMA}.vendor_stock_raw
                          (file_id, row_num, raw_sku, raw_qty, raw_payload, parsed_ok, parse_error)
                        VALUES %s
                        ON CONFLICT (file_id, row_num) DO UPDATE
                          SET raw_sku = EXCLUDED.raw_sku,
                              raw_qty = EXCLUDED.raw_qty,
                              raw_payload = EXCLUDED.raw_payload,
                              parsed_ok = EXCLUDED.parsed_ok,
                              parse_error = EXCLUDED.parse_error;
                        """,
                        raw_rows2,
                        page_size=2000,
                    )

                if norm_rows2:
                    psycopg2.extras.execute_values(
                        cur,
                        f"""
                        INSERT INTO {SCHEMA}.inventory_normalized
                          (file_id, vendor_id, as_of_date, sku, qty, source_row_num)
                        VALUES %s
                        ON CONFLICT (file_id, sku) DO UPDATE
                          SET qty = EXCLUDED.qty,
                              source_row_num = EXCLUDED.source_row_num,
                              normalized_at = NOW();
                        """,
                        norm_rows2,
                        page_size=2000,
                    )

                cur.execute(
                    f"""
                    UPDATE {SCHEMA}.vendor_files
                    SET status = 'normalized',
                        processed_at = NOW(),
                        error_message = NULL
                    WHERE file_id = %s;
                    """,
                    (fid,),
                )

        conn.commit()

    print("Ingestion complete")
    print(f"  File:   {os.path.basename(file_path)}")
    print(f"  Hash:   {file_hash}")
    print(f"  as_of:  {as_of.isoformat()}")
    print(f"  Sheet:  {args.sheet}")
    print(f"  Marker rows detected (non-column-A images): {marker_rows}")
    print(f"  Brand marker mapping used: {row_to_brand}")
    for b in sorted(raw_by_brand.keys()):
        print(f"  {b}: raw={len(raw_by_brand[b])}, normalized={len(norm_by_brand[b])}")


if __name__ == "__main__":
    main()
